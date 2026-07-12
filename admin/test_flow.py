"""Acceptance for the Admin & Publishing flow (Step 8).

Malformed upload → blocked. Clean pair → preview green + AAA634 sample → publish → audit → rollback.
Needs the real workbooks in pipeline/data/ (git-ignored), so the whole module skips without them.
"""
import shutil
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
import flow  # noqa: E402

INC = flow.PIPE / "data/Incentive_Sheet_Apr.xlsx"
SP = flow.PIPE / "data/Final_SP_Summary_Apr_26_Field.xlsb"
pytestmark = pytest.mark.skipif(not (INC.exists() and SP.exists()),
                                reason="real workbooks not present (data is git-ignored)")


@pytest.fixture
def bad_incentive(tmp_path):
    """Copy the real incentive sheet and inject one out-of-range value (persistency_cm = 9.99, le=3)."""
    dst = tmp_path / "bad_incentive.xlsx"
    shutil.copy(INC, dst)
    wb = load_workbook(dst)
    ws = wb["DSE"]
    ws.cell(row=3, column=17, value=9.99)   # persistency_cm (0-based idx 16) → violates le=3
    wb.save(dst)
    return dst


def test_malformed_upload_is_blocked(bad_incentive):
    flow.save_upload(bad_incentive, SP, user="tester")
    rep = flow.validate()
    assert rep["ok"] is False
    assert rep["blocking"], "a malformed sheet must produce a blocking reason"
    assert rep["incentive"]["range_errors"] or rep["incentive"].get("structural_error")
    with pytest.raises(flow.FlowError):
        flow.publish(user="tester")           # publish refused while invalid/unreconciled


def test_clean_pair_previews_publishes_audits_rolls_back():
    flow.save_upload(INC, SP, user="tester")

    rep = flow.validate()
    assert rep["ok"] is True and not rep["blocking"]
    assert rep["coverage"]["both"] == 812

    pv = flow.preview()
    assert pv["reconcile"]["ok"] is True
    assert pv["reconcile"]["incentive"]["checked"] == 891 and pv["reconcile"]["incentive"]["mismatches"] == 0
    assert pv["reconcile"]["sp"]["mismatches"] == 0
    assert pv["dse_count"] == 1098
    assert abs(pv["sample"]["incentive_final"] - 5356.336356) < 0.01
    assert pv["sample"]["sp_eligible"] is False and pv["sample"]["sp_tier"] == "on_track"

    v1 = flow.publish(user="tester")
    assert v1["reconcile"]["incentive_ok"] and v1["reconcile"]["sp_ok"]
    audit = flow.audit()
    assert any(a["id"] == v1["id"] and a["is_live"] for a in audit)

    # a second publish becomes live, then rollback restores v1 as live
    flow.save_upload(INC, SP, user="tester")
    flow.preview()
    v2 = flow.publish(user="tester")
    assert v2["id"] != v1["id"]
    assert any(a["id"] == v2["id"] and a["is_live"] for a in flow.audit())

    flow.rollback(v1["id"], user="tester")
    live = [a for a in flow.audit() if a["is_live"]]
    assert len(live) == 1 and live[0]["id"] == v1["id"]
