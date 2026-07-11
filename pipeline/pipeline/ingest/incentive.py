"""
Ingest the monthly Incentive workbook (ABSLI 'DSE' sheet) into the canonical input frame.

Reads cached values (data_only) since the sheet's inputs arrive via external VLOOKUPs. Maps the
ABSLI 39-column layout to clean canonical names. `sheet_final_amount` (column AD) is carried
through only so the reconciliation test can check the engine against the sheet's own output.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Optional

import pandas as pd
from openpyxl import load_workbook

# 0-based column indices in the 'DSE' sheet (header on row 2, data from row 3).
COLS = {
    "agent_code": 0, "employee_code": 1, "branch_code": 2, "status": 3,
    "name": 4, "grade": 5, "target_monthly": 6, "wfyp_others": 7,
    "wfyp_ulip_gap": 8, "ulip_fyp": 12, "persistency_cm": 16, "persistency_lm": 17,
    "nop_count": 24, "sheet_final_amount": 29, "sm_code": 32, "rsm_code": 34, "zsm_code": 36,
}
_PIFA_STATUS_COL = 27          # "25% Incentive Hold"
PIFA_HOLD_TRIGGER = "Not Achieved"
_CODE_COLS = ["agent_code", "employee_code", "branch_code", "status",
              "name", "grade", "sm_code", "rsm_code", "zsm_code"]
_NUMERIC_COLS = ["target_monthly", "wfyp_others", "wfyp_ulip_gap", "ulip_fyp",
                 "persistency_cm", "persistency_lm", "nop_count", "sheet_final_amount"]


def _code(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return s or None


def load_incentive(path: str | Path, sheet: str = "DSE", header_row: int = 2) -> pd.DataFrame:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows: list[dict] = []
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not r:
            continue
        agent = r[COLS["agent_code"]] if len(r) > COLS["agent_code"] else None
        if agent is None or str(agent).strip() == "":
            continue  # skip blank / total rows
        rec = {name: (r[idx] if len(r) > idx else None)
               for name, idx in COLS.items() if not name.startswith("_")}
        pifa_raw = r[_PIFA_STATUS_COL] if len(r) > _PIFA_STATUS_COL else None
        rec["pifa_ytd_met"] = str(pifa_raw).strip() != PIFA_HOLD_TRIGGER
        rec["pifa_status_raw"] = pifa_raw
        rows.append(rec)
    wb.close()

    df = pd.DataFrame(rows)
    for c in _CODE_COLS:
        if c in df.columns:
            df[c] = df[c].map(_code)
    for c in _NUMERIC_COLS:                    # '#N/A' and other non-numerics → NaN
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df
