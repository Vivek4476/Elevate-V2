"""
Generate reference templates that MIRROR THE REAL IMPORT FORMAT.

The monthly upload is the incentive team's actual sheet (the 'DSE' tab) and the real SP dashboard.
So the template is not an idealised schema — it reproduces those exact sheets: same columns, same
order, same header row position, read straight from the real files. Two illustrative sample rows are
included with identifier columns anonymised (so no real DSE's earnings circulate). Each template also
carries a Data Dictionary (marking which columns are pipeline INPUTS vs recomputed) and a Rules tab.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ARIAL = "Arial"
RED = "C00000"
GREY = "F2F2F2"
YELLOW = "FFF2CC"      # INPUT columns ops must get right
BLUE = "DDEBF7"        # group-header band
_thin = Side(style="thin", color="D9D9D9")
_B = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

# ── which real columns are pipeline INPUTS, and which identifiers to anonymise ──
INC_INPUTS = {"WFYP Target", "WFYP (Others)", "WFYP (ULIP+GAP)", "FYP (ULIP+GAP)",
              "Persistency (CM) (ii)", "Persistency (LM) (i)", "NOP Count", "25% Incentive Hold"}
INC_COMPUTED = {"WFYP Ach%", "WFYP Payout %", "WFYP Payout (Others)(A)", "(ULIP+GAP) Grid %",
                "(ULIP+GAP) Payout (B)", "Total Payout (A+B)", "Persistancy Growth (i-ii)",
                "Persistency Slab 1", "Persistency Slab 2", "Max", "Persistency Booster",
                "Payout Post Persistency", "NOP Multiplier", "NOP Payout", "Hold Amount",
                "Final Amount", "Already Paid", "Diff", "DSE Benchmark"}
INC_ANON = {"Agent Code", "Employee Code", "Agent Name",
            "SM Code", "DASM Code", "RSM Code", "Sr. RSM Code", "ZSM Code"}
INC_MEANING = {
    "WFYP Target": "Monthly WFYP target (₹)",
    "WFYP (Others)": "Non-ULIP WFYP for the month (₹) — may be negative (surrenders/clawbacks)",
    "WFYP (ULIP+GAP)": "ULIP+GAP WFYP for the month (₹)",
    "FYP (ULIP+GAP)": "ULIP+GAP FYP for the month (₹) — drives the ULIP slab",
    "Persistency (CM) (ii)": "13-month persistency, current month (0–1)",
    "Persistency (LM) (i)": "13-month persistency, last month (0–1)",
    "NOP Count": "Policies (NOP) for the month",
    "25% Incentive Hold": "'Not Achieved' ⇒ YTD PIFA not met → 25% held",
    "Agent Code": "Unique DSE agent code",
    "Employee Code": "HR employee code — JOIN KEY to SP 'DSE ID'",
    "Final Amount": "Final monthly incentive — the pipeline recomputes this and reconciles to it",
}

SP_INPUTS = {"WFYP YTD Target", "WFYP YTD Ach", "NOP YTD Target", "NOP YTD Ach", "Persistency % Overall"}
SP_COMPUTED = {"Ach %", "Achv Criteria Met", "WFYP WAS", "NOP WAS", "Overall Final WAS",
               "Reason for Promotion", "PIP WFYP Target", "PIP Ach", "PIP Remarks",
               "Actual WFYP", "Achv. %"}
SP_ANON = {"DSE ID", "DSE BO Code", "Employee Name",
           "SM Code", "SM Name", "RSM Code", "RSM", "ZSM Code", "ZSM"}
SP_MEANING = {
    "DSE ID": "HR employee code — JOIN KEY to Incentive 'Employee Code'",
    "WFYP YTD Target": "Rolling-12m WFYP target (₹)",
    "WFYP YTD Ach": "Rolling-12m WFYP achieved (₹) — may be negative (surrenders/clawbacks)",
    "NOP YTD Target": "Rolling-12m NOP target",
    "NOP YTD Ach": "Rolling-12m NOP achieved",
    "Persistency % Overall": "Overall persistency (0–1)",
    "Overall Final WAS": "0.75·WFYP% + 0.25·NOP% — the pipeline recomputes & reconciles this",
}


def _title(cell, text):
    cell.value = text
    cell.font = Font(name=ARIAL, bold=True, size=13, color=RED)


def _cell(cell, text, *, bold=False, italic=False, fill=None, color="000000", center=False):
    cell.value = text
    cell.font = Font(name=ARIAL, size=10, bold=bold, italic=italic, color=color)
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)
    cell.border = _B
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)


def _role(header, inputs, computed):
    if header in inputs:
        return "INPUT"
    if header in computed:
        return "computed"
    return "context"


def _anonymise(headers, row, sample_idx, anon):
    out = list(row)
    for i, h in enumerate(headers):
        if h in anon:
            out[i] = "Sample Name" if "Name" in h else f"SAMPLE{sample_idx + 1}"
    return out


def _read_incentive(path):
    ws = load_workbook(path, read_only=True, data_only=True)["DSE"]
    rows = list(ws.iter_rows(min_row=2, max_row=4, values_only=True))
    return list(rows[0]), [list(r) for r in rows[1:3]]


def _read_sp(path):
    raw = pd.read_excel(path, engine="pyxlsb", header=None, nrows=4)
    group = ["" if pd.isna(x) else str(x) for x in raw.iloc[0].tolist()]
    headers = ["" if pd.isna(x) else str(x) for x in raw.iloc[1].tolist()]
    samples = [["" if pd.isna(x) else x for x in raw.iloc[i].tolist()] for i in (2, 3)]
    return group, headers, samples


def _rules_lines(config):
    ic, sp = config["incentive"], config["sp"]
    p = lambda x: f"{x*100:g}%"
    L = ["MONTHLY INCENTIVE — how Final Amount is built", "",
         "Non-ULIP grid % (by overall WFYP achievement %):"]
    L += [f"   ≥ {p(b['from'])}  →  {p(b['pct'])}" for b in ic["wfyp_grid"]]
    L += ["", "ULIP grid % (by absolute ULIP FYP ₹):"]
    L += [f"   ≥ ₹{int(b['from']):,}  →  {p(b['pct'])}" for b in ic["ulip_slabs"]]
    L += [f"   (applies only when achievement ≥ {p(ic['ulip_gate_min_ach'])} of monthly target)",
          "", "Persistency multiplier (by current-month persistency):"]
    for b in ic["persistency"]["level_bands"]:
        L.append(f"   ≥ {p(b['from'])}  →  " + ("= persistency value" if b["mult"] == "equal_cm" else p(b["mult"])))
    imp = ic["persistency"]["improvement"]
    L += [f"   Improvement: if ≥{p(imp['min_level'])} AND up ≥{p(imp['min_growth'])} vs last month → {p(imp['mult'])} (max of the two)",
          "", "NOP multiplier (by policies):"]
    L += [f"   ≥ {int(b['from'])}  →  {p(b['mult'])}" for b in ic["nop_bands"]]
    L += ["", f"PIFA: {p(ic['pifa_hold_pct'])} held if YTD PIFA not achieved (YTD, not monthly).",
          "", "─" * 44, "", "SALES PROGRESSION (rolling 12 months)", "",
          f"Overall WAS = {sp['was_weights']['wfyp']:g}×WFYP ach% + {sp['was_weights']['nop']:g}×NOP ach%"]
    g = sp["gates"]
    L.append(f"Gates: WFYP ach ≥ {p(g['wfyp_ach_min'])} · NOP ach ≥ {p(g['nop_ach_min'])} · "
             f"WAS > {p(g['was_min'])} · Persistency ≥ {p(g['persistency_min'])}")
    return L


def _write_template(out_path, title, data_sheet, group_row, headers, samples, inputs, computed, anon, meaning, config):
    wb = Workbook()
    ws = wb.active
    ws.title = data_sheet
    hdr_row = 2                       # matches the real sheets: headers on row 2, data from row 3
    if group_row:
        for c, g in enumerate(group_row, start=1):
            if g:
                _cell(ws.cell(row=1, column=c), g, bold=True, fill=BLUE, center=True)
    else:
        _title(ws.cell(row=1, column=1), f"{title}  ·  headers on row 2, data from row 3")
    for c, h in enumerate(headers, start=1):
        role = _role(h, inputs, computed)
        _cell(ws.cell(row=hdr_row, column=c), h, bold=True, center=True,
              fill=(YELLOW if role == "INPUT" else GREY),
              color=("9C5700" if role == "INPUT" else "000000"))
        ws.column_dimensions[ws.cell(row=hdr_row, column=c).column_letter].width = max(12, min(30, len(str(h)) + 2))
    for s, srow in enumerate(samples):
        an = _anonymise(headers, srow, s, anon)
        for c, v in enumerate(an, start=1):
            _cell(ws.cell(row=hdr_row + 1 + s, column=c), "" if v is None else v)
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)

    # Data Dictionary
    wd = wb.create_sheet("Data Dictionary")
    _title(wd.cell(row=1, column=1), f"{title} — Data Dictionary")
    _cell(wd.cell(row=2, column=1),
          "This tab mirrors the exact monthly import format. Yellow = INPUT (must be accurate); "
          "'computed' columns are recomputed by the pipeline and used for reconciliation.", italic=True)
    for c, h in enumerate(["#", "Column (exact header)", "Meaning", "Role"], start=1):
        _cell(wd.cell(row=3, column=c), h, bold=True, fill=GREY, center=True)
    for i, h in enumerate(headers):
        role = _role(h, inputs, computed)
        _cell(wd.cell(row=4 + i, column=1), ws.cell(row=hdr_row, column=i + 1).column_letter, center=True)
        _cell(wd.cell(row=4 + i, column=2), h, bold=True)
        _cell(wd.cell(row=4 + i, column=3), meaning.get(h, {"INPUT": "pipeline input",
              "computed": "recomputed & reconciled by the pipeline", "context": "identifier / reference"}[role]))
        _cell(wd.cell(row=4 + i, column=4), role, center=True,
              fill=(YELLOW if role == "INPUT" else None), color=("9C5700" if role == "INPUT" else "000000"))
    for col, w in zip("ABCD", (6, 30, 52, 12)):
        wd.column_dimensions[col].width = w

    # Rules
    wr = wb.create_sheet("Rules")
    _title(wr.cell(row=1, column=1), "Incentive & SP rules (reference)")
    for i, line in enumerate(_rules_lines(config), start=3):
        _cell(wr.cell(row=i, column=1), line, bold=(line.isupper() and bool(line.strip())))
    wr.column_dimensions["A"].width = 92

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def build_all(config, out_dir, inc_wb, sp_wb):
    inc_headers, inc_samples = _read_incentive(inc_wb)
    sp_group, sp_headers, sp_samples = _read_sp(sp_wb)
    return [
        _write_template(Path(out_dir) / "Incentive_Import_Template.xlsx",
                        "Incentive (monthly) — 'DSE' import format", "DSE", None,
                        inc_headers, inc_samples, INC_INPUTS, INC_COMPUTED, INC_ANON, INC_MEANING, config),
        _write_template(Path(out_dir) / "SP_Import_Template.xlsx",
                        "Sales Progression (rolling 12m) import format", "SP Dashboard", sp_group,
                        sp_headers, sp_samples, SP_INPUTS, SP_COMPUTED, SP_ANON, SP_MEANING, config),
    ]
